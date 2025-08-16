"""
Excel Export Keywords Library for Robot Framework
Provides keywords for exporting Jira issues to Excel format
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from robot.api.logger import info, warn, error
from robot.api.deco import keyword


class ExcelExporter:
    """
   Robot Framework keyword library for Excel export functionality
   """

    ROBOT_LIBRARY_SCOPE = 'GLOBAL'

    def __init__(self):
        self.workbook = None
        self.worksheet = None

    @keyword('Export Issues To Excel')
    def export_issues_to_excel(self, issues_list, file_path, sheet_name='Jira Issues'):
        """
      Export Jira issues to Excel file with basic fields

      Args:
          issues_list: List of Jira issues
          file_path: Output Excel file path
          sheet_name: Name of the Excel sheet
      """
        if not issues_list:
            warn("No issues to export")
            return False

        info(f"Exporting {len(issues_list)} issues to {file_path}")

        # Create workbook and worksheet
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name

        # Define basic headers
        headers = [
            'Issue Key', 'Summary', 'Status', 'Assignee', 'Reporter',
            'Created', 'Priority', 'Issue Type', 'Description'
        ]

        # Write headers
        self._write_headers(headers)

        # Write issue data
        for row_num, issue in enumerate(issues_list, start=2):
            self._write_issue_row_basic(row_num, issue)

        # Apply formatting
        self._apply_basic_formatting()

        # Save file
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        self.workbook.save(file_path)

        info(f"Successfully exported {len(issues_list)} issues to {file_path}")
        return True

    @keyword('Export Issues To Excel With All Fields')
    def export_issues_to_excel_with_all_fields(self, issues_list, file_path, sheet_name='Jira Issues Comprehensive'):
        """
      Export Jira issues to Excel file with comprehensive field set

      Args:
          issues_list: List of Jira issues
          file_path: Output Excel file path
          sheet_name: Name of the Excel sheet
      """
        if not issues_list:
            warn("No issues to export")
            return False

        info(f"Exporting {len(issues_list)} issues with all fields to {file_path}")

        # Create workbook and worksheet
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name

        # Define comprehensive headers
        headers = [
            'Issue Key', 'Summary', 'Status', 'Assignee', 'Reporter',
            'Created', 'Updated', 'Priority', 'Issue Type', 'Project',
            'Components', 'Labels', 'Fix Versions', 'Due Date', 'Description'
        ]

        # Write headers
        self._write_headers(headers)

        # Write issue data
        for row_num, issue in enumerate(issues_list, start=2):
            self._write_issue_row_comprehensive(row_num, issue)

        # Apply formatting
        self._apply_comprehensive_formatting()

        # Save file
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        self.workbook.save(file_path)

        info(f"Successfully exported {len(issues_list)} issues with all fields to {file_path}")
        return True

    @keyword('Create Excel Report With Statistics')
    def create_excel_report_with_statistics(self, issues_list, file_path):
        """
      Create Excel report with issues data and statistics

      Args:
          issues_list: List of Jira issues
          file_path: Output Excel file path
      """
        if not issues_list:
            warn("No issues to create report")
            return False

        info(f"Creating Excel report with statistics for {len(issues_list)} issues")

        # Create workbook
        self.workbook = openpyxl.Workbook()

        # Remove default sheet
        self.workbook.remove(self.workbook.active)

        # Create issues sheet
        self._create_issues_sheet(issues_list)

        # Create statistics sheet
        self._create_statistics_sheet(issues_list)

        # Create summary sheet
        self._create_summary_sheet(issues_list)

        # Save file
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        self.workbook.save(file_path)

        info(f"Successfully created Excel report at {file_path}")
        return True

    def _write_headers(self, headers):
        """Write header row to worksheet"""
        for col_num, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def _write_issue_row_basic(self, row_num, issue):
        """Write basic issue data to row"""
        fields = issue.get('fields', {})

        # Extract basic field values
        values = [
            issue.get('key', ''),
            fields.get('summary', ''),
            fields.get('status', {}).get('name', '') if fields.get('status') else '',
            fields.get('assignee', {}).get('displayName', '') if fields.get('assignee') else 'Unassigned',
            fields.get('reporter', {}).get('displayName', '') if fields.get('reporter') else '',
            self._format_date(fields.get('created', '')),
            fields.get('priority', {}).get('name', '') if fields.get('priority') else '',
            fields.get('issuetype', {}).get('name', '') if fields.get('issuetype') else '',
            self._extract_description(fields.get('description', ''))
        ]

        # Write values to cells
        for col_num, value in enumerate(values, start=1):
            cell = self.worksheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def _write_issue_row_comprehensive(self, row_num, issue):
        """Write comprehensive issue data to row"""
        fields = issue.get('fields', {})

        # Extract comprehensive field values
        values = [
            issue.get('key', ''),
            fields.get('summary', ''),
            fields.get('status', {}).get('name', '') if fields.get('status') else '',
            fields.get('assignee', {}).get('displayName', '') if fields.get('assignee') else 'Unassigned',
            fields.get('reporter', {}).get('displayName', '') if fields.get('reporter') else '',
            self._format_date(fields.get('created', '')),
            self._format_date(fields.get('updated', '')),
            fields.get('priority', {}).get('name', '') if fields.get('priority') else '',
            fields.get('issuetype', {}).get('name', '') if fields.get('issuetype') else '',
            fields.get('project', {}).get('name', '') if fields.get('project') else '',
            self._extract_components(fields.get('components', [])),
            self._extract_labels(fields.get('labels', [])),
            self._extract_fix_versions(fields.get('fixVersions', [])),
            self._format_date(fields.get('duedate', '')),
            self._extract_description(fields.get('description', ''))
        ]

        # Write values to cells
        for col_num, value in enumerate(values, start=1):
            cell = self.worksheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def _apply_basic_formatting(self):
        """Apply basic formatting to worksheet"""
        # Auto-fit columns
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set column width with reasonable limits
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        self.worksheet.freeze_panes = 'A2'

    def _apply_comprehensive_formatting(self):
        """Apply comprehensive formatting to worksheet"""
        # Set specific column widths
        column_widths = {
            'A': 15,  # Issue Key
            'B': 40,  # Summary
            'C': 15,  # Status
            'D': 20,  # Assignee
            'E': 20,  # Reporter
            'F': 20,  # Created
            'G': 20,  # Updated
            'H': 12,  # Priority
            'I': 15,  # Issue Type
            'J': 20,  # Project
            'K': 25,  # Components
            'L': 25,  # Labels
            'M': 20,  # Fix Versions
            'N': 15,  # Due Date
            'O': 50  # Description
        }

        for column, width in column_widths.items():
            self.worksheet.column_dimensions[column].width = width

        # Freeze header row
        self.worksheet.freeze_panes = 'A2'

        # Set row height for better readability
        for row in range(2, self.worksheet.max_row + 1):
            self.worksheet.row_dimensions[row].height = 30

    def _create_issues_sheet(self, issues_list):
        """Create issues data sheet"""
        self.worksheet = self.workbook.create_sheet(title="Issues")

        headers = [
            'Issue Key', 'Summary', 'Status', 'Assignee', 'Reporter',
            'Created', 'Updated', 'Priority', 'Issue Type', 'Project', 'Description'
        ]

        self._write_headers(headers)

        for row_num, issue in enumerate(issues_list, start=2):
            self._write_issue_row_basic(row_num, issue)

        self._apply_basic_formatting()

    def _create_statistics_sheet(self, issues_list):
        """Create statistics sheet"""
        stats_sheet = self.workbook.create_sheet(title="Statistics")

        # Calculate statistics
        stats = self._calculate_statistics(issues_list)

        # Write statistics
        row = 1
        stats_sheet.cell(row=row, column=1, value="Jira Issues Statistics")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True, size=16)
        row += 2

        # Total issues
        stats_sheet.cell(row=row, column=1, value="Total Issues:")
        stats_sheet.cell(row=row, column=2, value=stats['total_issues'])
        row += 1

        # Status breakdown
        stats_sheet.cell(row=row, column=1, value="Status Breakdown:")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for status, count in stats['status_breakdown'].items():
            stats_sheet.cell(row=row, column=2, value=status)
            stats_sheet.cell(row=row, column=3, value=count)
            row += 1

        row += 1

        # Priority breakdown
        stats_sheet.cell(row=row, column=1, value="Priority Breakdown:")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for priority, count in stats['priority_breakdown'].items():
            stats_sheet.cell(row=row, column=2, value=priority)
            stats_sheet.cell(row=row, column=3, value=count)
            row += 1

        # Format statistics sheet
        for column in stats_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 30)
            stats_sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_summary_sheet(self, issues_list):
        """Create summary sheet"""
        summary_sheet = self.workbook.create_sheet(title="Summary")

        # Report header
        summary_sheet.cell(row=1, column=1, value="Jira Issues Export Report")
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=18)

        summary_sheet.cell(row=3, column=1, value="Export Date:")
        summary_sheet.cell(row=3, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        summary_sheet.cell(row=4, column=1, value="Total Issues Exported:")
        summary_sheet.cell(row=4, column=2, value=len(issues_list))

        # Format summary sheet
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 30

    def _calculate_statistics(self, issues_list):
        """Calculate statistics from issues list"""
        stats = {
            'total_issues': len(issues_list),
            'status_breakdown': {},
            'priority_breakdown': {},
            'assignee_breakdown': {}
        }

        for issue in issues_list:
            fields = issue.get('fields', {})

            # Status breakdown
            status = fields.get('status', {}).get('name', 'Unknown') if fields.get('status') else 'Unknown'
            stats['status_breakdown'][status] = stats['status_breakdown'].get(status, 0) + 1

            # Priority breakdown
            priority = fields.get('priority', {}).get('name', 'Unknown') if fields.get('priority') else 'Unknown'
            stats['priority_breakdown'][priority] = stats['priority_breakdown'].get(priority, 0) + 1

            # Assignee breakdown
            assignee = fields.get('assignee', {}).get('displayName', 'Unassigned') if fields.get(
                'assignee') else 'Unassigned'
            stats['assignee_breakdown'][assignee] = stats['assignee_breakdown'].get(assignee, 0) + 1

        return stats

    def _format_date(self, date_string):
        """Format date string for Excel"""
        if not date_string:
            return ''

        try:
            # Parse ISO format date
            if 'T' in date_string:
                dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            else:
                return date_string
        except:
            return date_string

    def _extract_description(self, description):
        """Extract description text"""
        if not description:
            return ''

        if isinstance(description, dict):
            # Handle ADF format
            return self._extract_text_from_adf(description)
        else:
            return str(description)

    def _extract_text_from_adf(self, adf_content):
        """Extract plain text from Atlassian Document Format (ADF)"""
        if not adf_content or not isinstance(adf_content, dict):
            return ''

        text = ''
        content = adf_content.get('content', [])

        for item in content:
            if item.get('type') == 'paragraph':
                paragraph_content = item.get('content', [])
                for text_item in paragraph_content:
                    if text_item.get('type') == 'text':
                        text += text_item.get('text', '') + ' '
            elif item.get('type') == 'text':
                text += item.get('text', '') + ' '

        return text.strip()

    def _extract_components(self, components):
        """Extract components list as string"""
        if not components:
            return ''

        component_names = [comp.get('name', '') for comp in components if comp.get('name')]
        return ', '.join(component_names)

    def _extract_labels(self, labels):
        """Extract labels list as string"""
        if not labels:
            return ''

        return ', '.join(labels)

    def _extract_fix_versions(self, fix_versions):
        """Extract fix versions list as string"""
        if not fix_versions:
            return ''

        version_names = [version.get('name', '') for version in fix_versions if version.get('name')]
        return ', '.join(version_names)
